import openpyxl

# Открываем файл Excel
file_path = "cleaned_file.xlsx"  # Замени на свой путь к файлу
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # Если у тебя только один лист в файле, то используем active

# Идем с конца листа, чтобы не сломать индексацию при удалении строк
for row in range(ws.max_row, 0, -1):
    # Проверяем, пусты ли все ячейки в строке
    if all([cell.value is None for cell in ws[row]]):
        ws.delete_rows(row, 1)  # Удаляем строку

# Сохраняем изменения в файл
wb.save("cleaned_file_no_empty_rows.xlsx")
print("Удаление пустых строк завершено, файл сохранен как cleaned_file_no_empty_rows.xlsx")
