import openpyxl
from openpyxl.styles import Font

# Открываем файл Excel
file_path = "supersliv.xlsx"  # Замени на свой путь к файлу
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # Если у тебя только один лист в файле, то используем active

# Пробегаем по всем ячейкам в колонках A и B
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
    for cell in row:
        # Проверяем, есть ли у ячейки текст и установлен ли шрифт
        if cell.value and cell.font and cell.font.color:
            # Проверяем, если цвет текста красный
            if cell.font.color.rgb == "FFFF0000":  # Красный цвет в формате RGB
                cell.value = None  # Очищаем содержимое ячейки

# Сохраняем изменения в файл
wb.save("cleaned_file.xlsx")
print("Удаление мусора завершено, файл сохранен как cleaned_file.xlsx")
